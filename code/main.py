from openpyxl import load_workbook
import datetime
import re

def parseWeek(sheet, teacher, schedule, week):
    teacherRange = teacher.get("range")
    typeOfLessonRegEx = re.compile("лаб|лек")
    for day in week.keys():
        currentIndex = 0
        for row in range(teacherRange[1], teacherRange[3] + 1):
            for col in range(day[0], day[2] + 1):
                cell = sheet.cell(row, col)
                if (cell.value == None):
                    continue
                if (row >= teacher.get("range")[1] and row <= teacher.get("range")[3]): # branch of lessons
                    for day in week.keys():
                        newRange = [day[0], day[1], day[2], teacher.get("range")[3]]
                        if (coordInRange(coord = cell.coordinate, range = newRange)):
                            # print(week.get(day))
                            if (schedule.get(teacher.get("name")) == None):
                                    schedule[teacher.get("name")] = {week.get(day): []}
                            try:
                                lessons = schedule[teacher.get("name")][week.get(day)]
                                lesson = lessons[currentIndex]
                                valueSet = False
                                # for lesson in lessons:
                                while (True):
                                    if (lesson.get("nameOfLesson") == None):
                                        lesson["nameOfLesson"] = cell.value
                                        valueSet = True
                                        break
                                    elif (col > newRange[0] and col < newRange[2] and re.search("\d{4}", str(prevValue)) != None):
                                        newName = lesson["nameOfLesson"] + " " + cell.value
                                        lesson["nameOfLesson"] = newName
                                        valueSet = True
                                        break
                                    if (lesson.get("office") == None):
                                        lesson["office"] = cell.value
                                        valueSet = True
                                        break
                                    if (lesson.get("groups") == None):
                                        lesson["groups"] = [cell.value]
                                        valueSet = True
                                        break
                                    elif (col == newRange[0] and re.search("\d{2}:\d{2}", str(cell.value)) == None and typeOfLessonRegEx.match(cell.value) == None):
                                        groups = lesson["groups"]
                                        groups.append(cell.value)
                                        lesson["groups"] = groups
                                        valueSet = True
                                        break
                                    if (lesson.get("type") == None):
                                        lesson["type"] = [cell.value]
                                        valueSet = True
                                        break
                                    elif (col == newRange[2] and re.search("\d{4}", str(cell.value)) == None):
                                        types = lesson["type"]
                                        types.append(cell.value)
                                        lesson["type"] = types
                                        valueSet = True
                                        break
                                    break
                                if (valueSet == False):
                                    time = cell.value
                                    print("append Time")
                                    if (isinstance(time, datetime.time)):
                                        strTime = time.strftime("%H:%M")
                                        lessons.append({"time": strTime})
                                        currentIndex += 1
                                    else:
                                        lessons.append({"time": time})
                                        currentIndex += 1
                                        # if (lessons["time"] != None and lessons["groups"] != None and lesson["office"] and lesson["nameOfLesson"] != None):
                                        #     lesson["nameOfLesson"] = lesson["nameOfLesson"] + cell.value
                                        #     break
                            except Exception as ex:
                                print("except")
                                print(ex)
                                time = cell.value
                                # print(time)
                                # print(isinstance(time, datetime.time))
                                if (isinstance(time, datetime.time)):
                                        strTime = time.strftime("%H:%M")
                                        # print({"time": strTime})
                                        schedule[teacher.get("name")][week.get(day)] = [{"time": strTime}]
                                else: 
                                    schedule[teacher.get("name")][week.get(day)] = [{"time": time}]
                            break
                    # print(cell.value)
                prevValue = cell.value

def getCharByColumn(column=1):
    charColumn = "A"
    counter = 1  
    while(counter != column):
        counter += 1
        charColumn = chr(ord(charColumn) + 1)
    return charColumn

def getRange(range):
    arrayOfCoordinates = []
    [firstPart, lastPart] = [getCharByColumn(range[0]) + str(range[1]), getCharByColumn(range[2]) + str(range[3])]
    tmpFirstPart = firstPart
    tmpNumber = range[1]
    while(tmpFirstPart != lastPart):
        arrayOfCoordinates.append(tmpFirstPart)
        if (firstPart[0] != lastPart[0]):
            tmpFirstPart = chr(ord(tmpFirstPart[0]) + 1) + str(tmpNumber)
        if (tmpFirstPart[0] == lastPart[0] and tmpFirstPart != lastPart):
            arrayOfCoordinates.append(tmpFirstPart)
            tmpNumber += 1
            tmpFirstPart = firstPart[0] + str(tmpNumber)
    arrayOfCoordinates.append(tmpFirstPart) # append last coordinate of range
    return arrayOfCoordinates

def coordInRange(coord="", range=""):
    arrayOfCoordinates = getRange(range)
    return coord in arrayOfCoordinates

def main():
    days = ["понедельник", "вторник", "среда", "четверг", "пятница", "суббота"]
    schedule = {}
    workbook = load_workbook('rasp.xlsx')
    sheets = workbook.worksheets
    for sheet in sheets:
        week = {}
        sortedMergedCells = sheet.merged_cells.sorted()
        print("max_row=", sheet.max_row+1)
        print("max_col=", sheet.max_column+1)
        teacher = {}
        for row in range(1,sheet.max_row+1):
            for col in range(1,sheet.max_column+1):
                if (len(week) == 6):
                    break
                cell = sheet.cell(row, col)
                if (cell.value == None):
                    continue
                if (isinstance(cell.value, str) and ("расписание" in cell.value.lower() or "семестр" in cell.value.lower())):
                    continue
                if (isinstance(cell.value, str) and cell.value.lower() in days):
                    for mergedCell in sortedMergedCells: 
                        if (coordInRange(coord = cell.coordinate, range = mergedCell.bounds)):
                            week[mergedCell.bounds] = cell.value # get dict of week with range coordinates
                            break
            if (len(week) == 6):
                break
        
        firstKey = None
        for key, val in week.items():
            firstKey = key

        for row in range(firstKey[1] + 1,sheet.max_row+1): # need to rewrite
            cell = sheet.cell(row, 1)
            teacher = None
            if (cell.value != None):
                # branch of teacher
                for mergedCell in sortedMergedCells: 
                    if (coordInRange(coord = cell.coordinate, range = mergedCell.bounds)):
                        teacher = {"name": cell.value, "range": mergedCell.bounds}
                        # print("###################################") 
                        # print(teacher)
                        break
                parseWeek(sheet, teacher, schedule, week)
                
            # if (col == 3 or col == 4):
            #     
            #     if (group == None and obj.get(cell.value) == None):
            #         obj[cell.value] = {}
            #         group = cell.value
            #         parseStart = True
            #         continue
            # if(parseStart):
            #     if (col == 1):
            #         obj[group][cell.value] = {}
            #         day = cell.value
            #     elif(col == 2):
            #         if (re.search("\d*\.\d*", cell.value) == None):
            #             obj[group][day][cell.value] = ""
            #             para = cell.value
            #     else:
            #         obj[group][day][para]=cell.value
                
            # print(cell.value)
            # print("#################################")
    # print(week)
    print(schedule)
    print(week)
            # print(row, col)
    # print(obj)
    with open('test.json', 'w', encoding='UTF-8') as f:
        f.write( re.sub(' +', ' ', str(schedule).replace("'", '"')))
    # sorteds = sheet.merged_cells.sorted()
    # print(sorteds)
    # print(sheet.cell(4,2).value)
    # print(sheet.cell(4,3).value)
    # print(sheet.cell(4,4).value)

if __name__ == "__main__":
    main()